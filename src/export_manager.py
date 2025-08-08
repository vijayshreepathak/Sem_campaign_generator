"""
Export Manager Module - Fixed Version
Handles exporting of all campaign deliverables to various formats.
"""

import pandas as pd
import os
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime

class ExportManager:
    """Manages export of campaign deliverables."""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.output_dir = config['output_settings']['directory']
        self.file_formats = config['output_settings']['file_formats']
        
        # Ensure output directory exists
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
    
    def export_all_deliverables(self, **datasets) -> List[str]:
        """
        Export all campaign deliverables.
        
        Args:
            **datasets: Named datasets to export
            
        Returns:
            List of exported file paths
        """
        exported_files = []
        
        # Export each dataset
        if 'raw_keywords' in datasets:
            files = self._export_keywords_raw(datasets['raw_keywords'])
            exported_files.extend(files)
        
        if 'processed_keywords' in datasets:
            files = self._export_master_keyword_list(datasets['processed_keywords'])
            exported_files.extend(files)
        
        if 'search_adgroups' in datasets:
            files = self._export_search_adgroups(datasets['search_adgroups'])
            exported_files.extend(files)
        
        if 'pmax_themes' in datasets:
            files = self._export_pmax_themes(datasets['pmax_themes'])
            exported_files.extend(files)
        
        if 'shopping_bids' in datasets:
            files = self._export_shopping_bids(datasets['shopping_bids'])
            exported_files.extend(files)
        
        # Export summary report
        summary_file = self._export_campaign_summary(**datasets)
        exported_files.append(summary_file)
        
        return exported_files
    
    def _export_keywords_raw(self, df: pd.DataFrame) -> List[str]:
        """Export raw keywords data."""
        files = []
        base_filename = "keywords_raw"
        
        for format_type in self.file_formats:
            if format_type == 'xlsx':
                filepath = os.path.join(self.output_dir, f"{base_filename}.xlsx")
                df.to_excel(filepath, index=False)
            elif format_type == 'csv':
                filepath = os.path.join(self.output_dir, f"{base_filename}.csv")
                df.to_csv(filepath, index=False)
            
            files.append(filepath)
        
        return files
    
    def _export_master_keyword_list(self, df: pd.DataFrame) -> List[str]:
        """Export processed master keyword list."""
        files = []
        base_filename = "master_keyword_list"
        
        # Reorder columns for better readability
        column_order = [
            'keyword', 'theme', 'search_intent', 'priority',
            'estimated_volume', 'estimated_competition', 
            'estimated_cpc_low', 'estimated_cpc_high',
            'composite_score', 'performance_rank', 'source'
        ]
        
        # Keep only existing columns
        existing_columns = [col for col in column_order if col in df.columns]
        export_df = df[existing_columns].copy()
        
        for format_type in self.file_formats:
            if format_type == 'xlsx':
                filepath = os.path.join(self.output_dir, f"{base_filename}.xlsx")
                self._export_simple_excel(export_df, filepath)
            elif format_type == 'csv':
                filepath = os.path.join(self.output_dir, f"{base_filename}.csv")
                export_df.to_csv(filepath, index=False)
            
            files.append(filepath)
        
        return files
    
    def _export_search_adgroups(self, df: pd.DataFrame) -> List[str]:
        """Export search campaign ad groups."""
        files = []
        base_filename = "Keyword_AdGroups"
        
        # Reorder columns for deliverable format
        column_order = [
            'ad_group', 'keyword', 'match_types', 
            'suggested_cpc_low', 'suggested_cpc_high',
            'estimated_volume', 'competition', 'search_intent',
            'composite_score', 'priority',
            'expected_monthly_clicks', 'expected_monthly_conversions',
            'theme_monthly_budget', 'source'
        ]
        
        existing_columns = [col for col in column_order if col in df.columns]
        export_df = df[existing_columns].copy()
        
        # Round numeric columns
        numeric_columns = [
            'suggested_cpc_low', 'suggested_cpc_high', 'composite_score',
            'expected_monthly_clicks', 'expected_monthly_conversions', 'theme_monthly_budget'
        ]
        
        for col in numeric_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(2)
        
        for format_type in self.file_formats:
            if format_type == 'xlsx':
                filepath = os.path.join(self.output_dir, f"{base_filename}.xlsx")
                self._export_simple_excel(export_df, filepath)
            elif format_type == 'csv':
                filepath = os.path.join(self.output_dir, f"{base_filename}.csv")
                export_df.to_csv(filepath, index=False)
            
            files.append(filepath)
        
        return files
    
    def _export_pmax_themes(self, df: pd.DataFrame) -> List[str]:
        """Export Performance Max themes."""
        files = []
        base_filename = "PMax_Themes"
        
        # Reorder columns
        column_order = [
            'theme_type', 'theme_name', 'description',
            'target_keywords', 'primary_intent', 'audience_signals',
            'monthly_budget', 'expected_impressions', 'expected_clicks', 
            'expected_conversions', 'target_roas', 'asset_requirements'
        ]
        
        existing_columns = [col for col in column_order if col in df.columns]
        export_df = df[existing_columns].copy()
        
        # Format target_keywords as string if it's a list
        if 'target_keywords' in export_df.columns:
            export_df['target_keywords'] = export_df['target_keywords'].apply(
                lambda x: ', '.join(x) if isinstance(x, list) else str(x)
            )
        
        # Round numeric columns
        numeric_columns = ['monthly_budget', 'expected_impressions', 'expected_clicks', 'expected_conversions', 'target_roas']
        for col in numeric_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(2)
        
        for format_type in self.file_formats:
            if format_type == 'xlsx':
                filepath = os.path.join(self.output_dir, f"{base_filename}.xlsx")
                self._export_simple_excel(export_df, filepath)
            elif format_type == 'csv':
                filepath = os.path.join(self.output_dir, f"{base_filename}.csv")
                export_df.to_csv(filepath, index=False)
            
            files.append(filepath)
        
        return files
    
    def _export_shopping_bids(self, df: pd.DataFrame) -> List[str]:
        """Export shopping campaign bids."""
        files = []
        base_filename = "Shopping_CPC_Bids"
        
        # Reorder columns
        column_order = [
            'product_category', 'category_description', 'priority_level',
            'avg_monthly_volume', 'competition_level',
            'top_of_page_bid_low', 'top_of_page_bid_high',
            'target_cpc_formula', 'suggested_cpc', 'margin_factor',
            'monthly_budget', 'expected_monthly_clicks', 'expected_monthly_conversions',
            'bidding_strategy'
        ]
        
        existing_columns = [col for col in column_order if col in df.columns]
        export_df = df[existing_columns].copy()
        
        # Round numeric columns
        numeric_columns = [
            'avg_monthly_volume', 'top_of_page_bid_low', 'top_of_page_bid_high',
            'target_cpc_formula', 'suggested_cpc', 'margin_factor', 'monthly_budget',
            'expected_monthly_clicks', 'expected_monthly_conversions'
        ]
        
        for col in numeric_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(2)
        
        for format_type in self.file_formats:
            if format_type == 'xlsx':
                filepath = os.path.join(self.output_dir, f"{base_filename}.xlsx")
                self._export_simple_excel(export_df, filepath)
            elif format_type == 'csv':
                filepath = os.path.join(self.output_dir, f"{base_filename}.csv")
                export_df.to_csv(filepath, index=False)
            
            files.append(filepath)
        
        return files
    
    def _export_campaign_summary(self, **datasets) -> str:
        """Export comprehensive campaign summary."""
        
        summary_data = []
        
        # Project information
        summary_data.append("# SEM Campaign Summary Report")
        summary_data.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary_data.append(f"Author: {self.config.get('project', {}).get('author', 'Vijayshree Vaibhav')}")
        summary_data.append("")
        
        # Configuration summary
        summary_data.append("## Campaign Configuration")
        summary_data.append(f"Brand: {self.config['brand']['name']}")
        summary_data.append(f"Competitor: {self.config['competitor']['name']}")
        summary_data.append(f"Service Locations: {', '.join(self.config['service_locations'])}")
        summary_data.append("")
        
        # Budget allocation
        summary_data.append("## Budget Allocation")
        budgets = self.config['budgets']
        summary_data.append(f"Search Campaigns: ${budgets['search_ads']:,}")
        summary_data.append(f"Shopping Campaigns: ${budgets['shopping_ads']:,}")
        summary_data.append(f"Performance Max: ${budgets['pmax_ads']:,}")
        summary_data.append(f"Total Monthly Budget: ${budgets['total_monthly']:,}")
        summary_data.append("")
        
        # Keywords summary
        if 'processed_keywords' in datasets:
            df = datasets['processed_keywords']
            summary_data.append("## Keywords Summary")
            summary_data.append(f"Total Processed Keywords: {len(df):,}")
            
            if 'theme' in df.columns:
                summary_data.append("\nKeywords by Theme:")
                theme_counts = df['theme'].value_counts()
                for theme, count in theme_counts.items():
                    summary_data.append(f"  • {theme}: {count:,}")
            
            if 'search_intent' in df.columns:
                summary_data.append("\nKeywords by Intent:")
                intent_counts = df['search_intent'].value_counts()
                for intent, count in intent_counts.items():
                    summary_data.append(f"  • {intent.title()}: {count:,}")
            
            summary_data.append("")
        
        # Campaign structure summary
        if 'search_adgroups' in datasets:
            df = datasets['search_adgroups']
            summary_data.append("## Search Campaign Structure")
            summary_data.append(f"Total Ad Groups: {df['ad_group'].nunique()}")
            summary_data.append(f"Total Keywords: {len(df):,}")
            
            ad_group_counts = df['ad_group'].value_counts()
            for ag, count in ad_group_counts.items():
                summary_data.append(f"  • {ag}: {count:,} keywords")
            summary_data.append("")
        
        if 'pmax_themes' in datasets:
            df = datasets['pmax_themes']
            summary_data.append("## Performance Max Themes")
            summary_data.append(f"Total Themes: {len(df)}")
            for _, row in df.iterrows():
                budget = row.get('monthly_budget', 0)
                summary_data.append(f"  • {row['theme_name']}: ${budget:,.0f} budget")
            summary_data.append("")
        
        if 'shopping_bids' in datasets:
            df = datasets['shopping_bids']
            summary_data.append("## Shopping Campaign Structure")
            summary_data.append(f"Product Categories: {len(df)}")
            for _, row in df.iterrows():
                budget = row.get('monthly_budget', 0)
                cpc = row.get('suggested_cpc', 0)
                summary_data.append(f"  • {row['product_category']}: ${cpc:.2f} CPC, ${budget:,.0f} budget")
            summary_data.append("")
        
        # Performance projections
        summary_data.append("## Expected Performance")
        
        total_clicks = 0
        total_conversions = 0
        
        if 'search_adgroups' in datasets:
            search_clicks = datasets['search_adgroups'].get('expected_monthly_clicks', pd.Series()).sum() if 'expected_monthly_clicks' in datasets['search_adgroups'].columns else 0
            search_conversions = datasets['search_adgroups'].get('expected_monthly_conversions', pd.Series()).sum() if 'expected_monthly_conversions' in datasets['search_adgroups'].columns else 0
            total_clicks += search_clicks
            total_conversions += search_conversions
            summary_data.append(f"Search Campaigns: {search_clicks:,.0f} clicks, {search_conversions:.1f} conversions")
        
        if 'pmax_themes' in datasets:
            pmax_clicks = datasets['pmax_themes'].get('expected_clicks', pd.Series()).sum() if 'expected_clicks' in datasets['pmax_themes'].columns else 0
            pmax_conversions = datasets['pmax_themes'].get('expected_conversions', pd.Series()).sum() if 'expected_conversions' in datasets['pmax_themes'].columns else 0
            total_clicks += pmax_clicks
            total_conversions += pmax_conversions
            summary_data.append(f"Performance Max: {pmax_clicks:,.0f} clicks, {pmax_conversions:.1f} conversions")
        
        if 'shopping_bids' in datasets:
            shopping_clicks = datasets['shopping_bids'].get('expected_monthly_clicks', pd.Series()).sum() if 'expected_monthly_clicks' in datasets['shopping_bids'].columns else 0
            shopping_conversions = datasets['shopping_bids'].get('expected_monthly_conversions', pd.Series()).sum() if 'expected_monthly_conversions' in datasets['shopping_bids'].columns else 0
            total_clicks += shopping_clicks
            total_conversions += shopping_conversions
            summary_data.append(f"Shopping Campaigns: {shopping_clicks:,.0f} clicks, {shopping_conversions:.1f} conversions")
        
        summary_data.append(f"\nTotal Projected: {total_clicks:,.0f} clicks, {total_conversions:.1f} conversions")
        
        # Write summary file
        summary_filepath = os.path.join(self.output_dir, "Campaign_Summary_Report.txt")
        with open(summary_filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(summary_data))
        
        return summary_filepath
    
    def _export_simple_excel(self, df: pd.DataFrame, filepath: str):
        """Export DataFrame to Excel without complex formatting to avoid errors."""
        try:
            # Simple Excel export without styling
            df.to_excel(filepath, index=False)
            
            # Optional: Add basic formatting if openpyxl is available
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = load_workbook(filepath)
                ws = wb.active
                
                # Format header row
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filepath)
                
            except ImportError:
                # openpyxl not available or formatting failed, use simple export
                pass
            except Exception:
                # Any formatting error, file is already saved with basic format
                pass
                
        except Exception as e:
            print(f"Warning: Excel export failed for {filepath}: {e}")
            # Fallback to CSV
            csv_filepath = filepath.replace('.xlsx', '.csv')
            df.to_csv(csv_filepath, index=False)
